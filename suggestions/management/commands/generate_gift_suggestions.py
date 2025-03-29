import os
import time
import json
import openpyxl
import re

from django.core.management.base import BaseCommand
from suggestions.models import (
    GiftCondition,
    GiftSuggestion,
    GENDER_CHOICES,
    AGE_GROUP_CHOICES,
    INTEREST_CHOICES,
    BUDGET_CHOICES,
)

from openai import OpenAI

# from dotenv import load_dotenv
# load_dotenv()

client = OpenAI(
   api_key='',
   base_url=''
)


# Excel setup
EXCEL_FILENAME = "gift_suggestions.xlsx"
if os.path.exists(EXCEL_FILENAME):
    workbook = openpyxl.load_workbook(EXCEL_FILENAME)
    sheet = workbook.active
    processed_ids = {int(row[0].value) for row in sheet.iter_rows(min_row=2)}
else:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(["Condition ID", "Gender", "Age Group", "Interest", "Budget", "Title", "Description", "Image URL", "Product URL"])
    processed_ids = set()

def get_label(choices, value):
    return dict(choices).get(value, value)

import re

def clean_reply(reply):
    reply = reply.strip()

    if reply.startswith("```json"):
        reply = reply[len("```json"):].strip()
    elif reply.startswith("```"):
        reply = reply[len("```"):].strip()
    if reply.endswith("```"):
        reply = reply[:-3].strip()

    if reply.startswith("{") and reply.endswith("}"):
        reply = f"[{reply}]"
    if not reply.startswith("["):
        reply = "[" + reply
    if not reply.endswith("]"):
        reply += "]"

    reply = re.sub(r'(?<!")(\b[\w\- ]+\b)(?=\s*:)', r'"\1"', reply)

    return reply



class Command(BaseCommand):
    help = "Generates gift suggestions using OpenAI and saves them to DB + Excel."

    def handle(self, *args, **kwargs):
        conditions = GiftCondition.objects.exclude(id__in=processed_ids)
        total = len(conditions)
        count = 0

        for cond in conditions:
            count += 1
            self.stdout.write(f"[{count}/{total}] Processing condition ID {cond.id}")

            prompt = (
                f"من یک هدیه‌ای می‌خوام برای یک {get_label(GENDER_CHOICES, cond.gender)} هستش که تقریبا سنش "
                f"{get_label(AGE_GROUP_CHOICES, cond.age_group)} هست. به {get_label(INTEREST_CHOICES, cond.interest)} "
                f"علاقه داره و می‌خوام که اندازه {get_label(BUDGET_CHOICES, cond.budget)} هزینه کنم. "
                f"برام مهمه که تو ایران بتونم پیشنهادات رو پیدا کنم، مثلا از دی‌جی‌کالا یا ترب یا با سلام.\n\n"
                f"و دقیقا در این قالب جواب بهم بده:\n"
                f"[\n"
                f"    {{product-title: ..., product-image: ..., product-description: ..., product-url: ...}},\n"
                f"    ...\n"
                f"]"
            )

            try:
                response = client.chat.completions.create(
                    model="gpt-4",
                    messages=[{"role": "user", "content": prompt}],
                    temperature=0.7
                )

                reply = response.choices[0].message.content

                try:
                    reply_clean = clean_reply(reply)
                    # reply_clean = reply.strip().removeprefix("```json").removesuffix("```").strip()
                    suggestions = json.loads(reply_clean.replace("'", '"'))
                except json.JSONDecodeError:
                    self.stdout.write(reply)
                    self.stdout.write(self.style.ERROR("❌ JSON parsing error"))
                    continue

                GiftSuggestion.objects.filter(condition=cond).delete()

                for s in suggestions:
                    GiftSuggestion.objects.create(
                        condition=cond,
                        title=s.get("product-title", "بدون عنوان"),
                        description=s.get("product-description", ""),
                        image_url=s.get("product-image", ""),
                        product_url=s.get("product-url", ""),
                        score=0
                    )

                    # Write to Excel
                    sheet.append([
                        cond.id,
                        get_label(GENDER_CHOICES, cond.gender),
                        get_label(AGE_GROUP_CHOICES, cond.age_group),
                        get_label(INTEREST_CHOICES, cond.interest),
                        get_label(BUDGET_CHOICES, cond.budget),
                        s.get("product-title", ""),
                        s.get("product-description", ""),
                        s.get("product-image", ""),
                        s.get("product-url", "")
                    ])

                workbook.save(EXCEL_FILENAME)
                self.stdout.write(self.style.SUCCESS("✅ Suggestions saved."))

                time.sleep(2)

            except Exception as e:
                self.stdout.write(self.style.ERROR(f"🔥 Error: {str(e)}"))
                workbook.save(EXCEL_FILENAME)  # Save before crash
                break
